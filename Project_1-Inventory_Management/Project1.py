import pulp
import openpyxl
import random


def readexcelfile(path, demand, manucost, sellingprice, products, daynames,numdays, numproducts):
    
    wb = openpyxl.load_workbook(filename = path)
    sheetnames = wb.sheetnames
    wsheet =wb[sheetnames[0]]
    for i in range (2, 2+numproducts):
        
        names.append(wsheet[i][0].value)
        sellingprice.append(float(wsheet[i][1].value))
        manucost.append(float(wsheet[i][2].value))

    for j in range (1, 1+numdays):
       daynames.append(wsheet[10][j].value)

    for i in range (11, 11+numproducts):
        temp=[]
        for j in range (1, 1+numdays):
            temp.append(float(wsheet[i][j].value))   
        demand.append(temp)

    
numproducts = 6
numdays = 7
path ="project1data.xlsx"
demand=[]
names=[]
daynames=[]
manucost=[]
sellingprice=[]



readexcelfile(path, demand, manucost, sellingprice,names, daynames,numdays, numproducts)

print ("demand is \n",demand)
print ("Products are \n",names)
print("Day names are \n",daynames)
print ("Manufacturing costs are \n",manucost)
print("selling prices are \n",sellingprice)
    
#Then instantiate a problem class, we’ll name it “My LP problem” and we’re looking for an optimal maximum so we use LpMaximize

my_lp_problem = pulp.LpProblem("My LP Problem", pulp.LpMaximize)

prod= pulp.LpVariable.dicts("production", ((i, j) for i in range(numproducts) for j in range(numdays)), lowBound=0, cat = 'Continuous')
invent= pulp.LpVariable.dicts("inventory", ((i, j) for i in range(numproducts) for j in range(numdays)), lowBound=0, cat = 'Continuous')
back= pulp.LpVariable.dicts("backorder", ((i, j) for i in range(numproducts) for j in range(numdays)), lowBound=0, cat = 'Continuous')
sold= pulp.LpVariable.dicts("sold", ((i, j) for i in range(numproducts) for j in range(numdays)), lowBound=0, cat = 'Continuous')

# obj and constr are added using += to the my_lp_problem class.

#Pulp assumes that the the objective function is alsyws given first.    

my_lp_problem += pulp.lpSum([sellingprice[i]*sold[i, j] for i in range(numproducts) for j in range(numdays)]) - pulp.lpSum([2000*prod[i, j] for i in range(numproducts) for j in range(numdays)]) - pulp.lpSum([manucost[i]*prod[i, j] for i in range(numproducts) for j in range(numdays)]) - pulp.lpSum([20*invent[i, j] for i in range(numproducts) for j in range(numdays)]) - pulp.lpSum([0.02*sellingprice[i]*back[i, j] for i in range(numproducts) for j in range(numdays)])

#constraints
for j in range (numdays):
    my_lp_problem += pulp.lpSum([prod[i, j] for i in range(numproducts)]) <=500
    
for i in range (numproducts):
    for j in range (numdays):
        my_lp_problem += sold[i, j] <= demand[i][j]      

for i in range (numproducts):
    for j in range (1,numdays):
        my_lp_problem += prod[i, j] + invent[i,j-1] + back[i,j] == sold[i,j] +  invent[i,j] + back[i,j-1]
        
for i in range (numproducts):
    my_lp_problem += prod[i, 0] + back[i,0] == sold[i,0] +  invent[i,0]
              
for i in range(numproducts):
    my_lp_problem += invent[i,numdays-1] ==0
    my_lp_problem += back[i,numdays-1] ==0
          
print (my_lp_problem)

status=my_lp_problem.solve()
if pulp.LpStatus[my_lp_problem.status] =='Infeasible':
    print ('INFEASIBLE ****************************************')
    
else:
    # To write out to an excel file
    
    
    wbr = openpyxl.Workbook()
    
    #Sheets start at 0 
    
    wbr.create_sheet(index = 0, title = "solution")
    sheetnames = wbr.sheetnames   
    
    wsheet = wbr[sheetnames[0]]
    
    #Writing to individual cells
    # Cells start at 1, 1
    
    #set column widths
    letters=['A','B','C','D','E','F','G','H','I','J','K','L']
    widths = [12, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8 ]
    for i in range (0,10):    
        wsheet.column_dimensions[letters[i]].width = widths[i]
    
    #I am using row to increase as I write out stuff and then columns.
    
    row =1 
    wsheet.cell(row,1).value ='Money Made'
    wsheet.cell(row,2).value =pulp.value(my_lp_problem.objective)
    row+=1
    
    for i in range (0, len(daynames)):
        wsheet.cell(row , i+2).value = daynames[i]
    row+=1
    wsheet.cell(row,1).value ='Production'
    row+=1
    for i in range (0,len(demand)):
        for j in range (0,len(demand[i])):
            wsheet.cell(row, j+2).value =prod[i,j].varValue
        wsheet.cell(row, j+3).value =names[i]
        row+=1
    
    wsheet.cell(row,1).value ='Inventory'
    row+=1
    for i in range (0,len(demand)):
        for j in range (0,len(demand[i])):
            wsheet.cell(row, j+2).value =invent[i,j].varValue
        wsheet.cell(row, j+3).value =names[i]
        row+=1
    
    wsheet.cell(row,1).value ='Backorder'
    row+=1
    for i in range (0,len(demand)):
        for j in range (0,len(demand[i])):
            wsheet.cell(row, j+2).value =back[i,j].varValue
        wsheet.cell(row, j+3).value =names[i]
        row+=1
    
    wsheet.cell(row,1).value ='Sold'
    row+=1
    for i in range (0,len(demand)):
        for j in range (0,len(demand[i])):
            wsheet.cell(row, j+2).value =sold[i,j].varValue
        wsheet.cell(row, j+3).value =names[i]
        row+=1
    
    wbr.save('readinoutputexcel.xlsx')
