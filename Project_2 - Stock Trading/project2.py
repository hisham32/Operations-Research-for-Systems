# -*- coding: utf-8 -*-
"""
Created on Wed Oct  5 20:25:42 2022

@author: u1400077
"""

import pulp
import openpyxl
import random
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

# Loading stock price data
stock = pd.read_excel('Stocks.xlsx')

# Processing data
numproducts = len(stock.columns.tolist()[2:])
numdays = len(stock)
names = stock.columns.tolist()[2:]
stockp = np.transpose(stock[names].values)
monthnames=stock['Month'].tolist()
daynames=stock['Day'].tolist()

# Visualizing stock prices
stocks=stock[names]
for i in stocks.columns.tolist():
    stocks[[i]].plot(xlabel='Trading Days', ylabel='Stock Price ($)').get_figure().savefig('X:/Documents/fig_'+i)

    
#Then instantiate a problem class, we’ll name it “My LP problem” and we’re looking for an optimal maximum so we use LpMaximize

my_lp_problem = pulp.LpProblem("My LP Problem", pulp.LpMaximize)

# decision variables
buy= pulp.LpVariable.dicts("buy", ((i, j) for i in range(numproducts) for j in range(numdays)), lowBound=0, cat = 'Continuous')
sold= pulp.LpVariable.dicts("sold", ((i, j) for i in range(numproducts) for j in range(numdays)), lowBound=0, cat = 'Continuous')
own= pulp.LpVariable.dicts("own", ((i,j) for i in range(numproducts) for j in range(numdays)), lowBound=0, cat = 'Continuous')
cash= pulp.LpVariable.dicts("cash", (j for j in range(numdays)), lowBound=0, cat = 'Continuous')

# obj and constr are added using += to the my_lp_problem class.

#Pulp assumes that the the objective function is alsyws given first.     
my_lp_problem += cash[numdays-1] 

#constraints     
for i in range (numproducts):
    for j in range (1,numdays):
        my_lp_problem += own[i,j-1] + buy[i,j] - sold[i,j] == own[i,j]
        
for i in range(numproducts):
    my_lp_problem += buy[i,0] == own[i,0] 
    
for i in range (numproducts):
    for j in range (1,numdays):
        my_lp_problem += sold[i, j] <= own[i,j-1] 
            
for j in range (1, numdays):
    my_lp_problem += 1.00008 * cash[j-1] + pulp.lpSum([sold[i,j] * stockp[i,j] for i in range(numproducts)]) - pulp.lpSum([buy[i,j] * stockp[i,j] for i in range(numproducts)])  == cash[j]    
        

my_lp_problem +=  10000000 - pulp.lpSum([buy[i,0] * stockp[i,0] for i in range(numproducts)])  == cash[0]   
    
for j in range (1,numdays):
    my_lp_problem += pulp.lpSum([buy[i,j] * stockp[i,j] for i in range(numproducts)]) <= cash[j-1]
              
for i in range(numproducts):
    my_lp_problem += buy[i,numdays-1] == 0
    my_lp_problem += sold[i,0] == 0 
          
print (my_lp_problem)

status=my_lp_problem.solve()
if pulp.LpStatus[my_lp_problem.status] =='Infeasible':
    print ('INFEASIBLE ****************************************')
    
else:
    # To write out to an excel file
    
    
    wbr = openpyxl.Workbook()
    
    #Sheets start at 0 
    
    wbr.create_sheet(index = 0, title = "Bought")
    wbr.create_sheet(index = 1, title = "Sold")
    wbr.create_sheet(index = 2, title = "Owned")
    wbr.create_sheet(index = 3, title = "Cash")
    sheetnames = wbr.sheetnames   
    
    wsheet = wbr[sheetnames[0]]
    
    #Writing to individual cells
    # Cells start at 1, 1
    
    #set column widths
    letters=['A','B','C','D','E','F','G','H','I','J','K']
    widths = [8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8]
    for i in range (0,9):    
        wsheet.column_dimensions[letters[i]].width = widths[i]
    
    #I am using row to increase as I write out stuff and then columns.
    
    row =1 
    
    for i in range (0, len(names)):
        wsheet.cell(row , i+4).value = names[i]
    wsheet.cell(row,1).value ='Bought'
    wsheet.cell(row , 2).value = 'Month'
    wsheet.cell(row , 3).value = 'Day'
    
    row+=1
    for j in range (0,numdays):
        wsheet.cell(row, 2).value =monthnames[j]
        wsheet.cell(row, 3).value =daynames[j]
        for i in range (0,len(names)):
            wsheet.cell(row, i+4).value =buy[i,j].varValue

        row+=1

    wsheet = wbr[sheetnames[1]]
    
    #Writing to individual cells
    # Cells start at 1, 1
    
    #set column widths
    letters=['A','B','C','D','E','F','G','H','I','J','K']
    widths = [8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8]
    for i in range (0,9):    
        wsheet.column_dimensions[letters[i]].width = widths[i]
    
    #I am using row to increase as I write out stuff and then columns.
    
    row =1 
    
    for i in range (0, len(names)):
        wsheet.cell(row , i+4).value = names[i]
    wsheet.cell(row,1).value ='Sold'
    wsheet.cell(row , 2).value = 'Month'
    wsheet.cell(row , 3).value = 'Day'
    #
    row+=1
    for j in range (0,numdays):
        wsheet.cell(row, 2).value =monthnames[j]
        wsheet.cell(row, 3).value =daynames[j]
        for i in range (0,len(names)):
            wsheet.cell(row, i+4).value =sold[i,j].varValue

        row+=1

    
    wsheet = wbr[sheetnames[2]]
    
    #Writing to individual cells
    # Cells start at 1, 1
    
    #set column widths
    letters=['A','B','C','D','E','F','G','H','I','J','K']
    widths = [8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8]
    for i in range (0,9):    
        wsheet.column_dimensions[letters[i]].width = widths[i]
    
    #I am using row to increase as I write out stuff and then columns.
    
    row =1 
    
    for i in range (0, len(names)):
        wsheet.cell(row , i+4).value = names[i]
    wsheet.cell(row,1).value ='Owned'
    wsheet.cell(row , 2).value = 'Month'
    wsheet.cell(row , 3).value = 'Day'
    
    row+=1
    for j in range (0,numdays):
        wsheet.cell(row, 2).value =monthnames[j]
        wsheet.cell(row, 3).value =daynames[j]
        for i in range (0,len(names)):
            wsheet.cell(row, i+4).value =own[i,j].varValue

        row+=1
    
    wsheet = wbr[sheetnames[3]]
    
    #Writing to individual cells
    # Cells start at 1, 1
    
    #set column widths
    letters=['A','B','C','D','E','F','G','H','I','J','K']
    widths = [8, 12, 8, 12, 8, 8, 8, 8, 8, 8, 8]
    for i in range (0,9):    
        wsheet.column_dimensions[letters[i]].width = widths[i]
    
    #I am using row to increase as I write out stuff and then columns.
    
    row =1 
    wsheet.cell(row,1).value ='Money Made'
    wsheet.cell(row,2).value =pulp.value(my_lp_problem.objective)
    row+=1
    
    
    wsheet.cell(row, 1).value ='Cash'
    wsheet.cell(row , 2).value = 'Month'
    wsheet.cell(row , 3).value = 'Day'
    wsheet.cell(row , 4).value = 'Cash ($)'
    row+=1
    for j in range (0,numdays):
        wsheet.cell(row, 2).value =monthnames[j]
        wsheet.cell(row, 3).value =daynames[j]
        wsheet.cell(row, 4).value = cash[j].varValue

        row+=1    
    
    wbr.save('readinoutputexcel2.xlsx')
