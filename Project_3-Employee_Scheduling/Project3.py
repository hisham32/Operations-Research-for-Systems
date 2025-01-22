import openpyxl
import pulp
import openpyxl


class Employee (object):
  def __init__(self,eid, pay, lowhours, highhours, cashier, stocking, customerservice, backroom, floorassociate, availability):
                self.eid = eid
                self.pay = pay
                self.lowhours = lowhours
                self.highhours = highhours
                self.cashier = cashier
                self.stocking = stocking
                self.customerservice = customerservice
                self.backroom = backroom
                self.floorassociate = floorassociate
                self.availability = availability
                
     
def readexcelfile(path, employees, cashier, stocking, customerservice, backroom, floorassociate):
    
    wb = openpyxl.load_workbook(filename = path)
    sheetnames = wb.sheetnames
    wsheet =wb[sheetnames[0]]
    print (wsheet.max_column, wsheet.max_row)
    
    for j in range (1,wsheet.max_column):
        eid = wsheet[1][j].value
        pay = float(wsheet[2][j].value)
        #print (pay)
        lowhours = int(wsheet[3][j].value)
        highhours = int(wsheet[4][j].value)
        cashier = int(wsheet[5][j].value)
        stocking = int(wsheet[6][j].value)
        customerservice = int(wsheet[7][j].value)
        backroom = int(wsheet[8][j].value)
        floorassociate= int(wsheet[9][j].value)
        availability=[]
        for i in range (11, wsheet.max_row+1):
            availability.append(int(wsheet[i][j].value))
        employees.append(Employee(eid, pay, lowhours, highhours, cashier, stocking, customerservice, backroom, floorassociate, availability))
    
    wsheet =wb[sheetnames[1]]
    for i in range (2,wsheet.max_row+1):
         rcashier.append(int(wsheet[i][1].value))
         rcustomerservice.append(int(wsheet[i][2].value))
         rstocking.append(int(wsheet[i][3].value))
         rbackroom.append(int(wsheet[i][4].value))
         rfloorassociate.append(int(wsheet[i][5].value))
         

path ="Project3data.xlsx"


employees=[]
rcashier=[]
rstocking=[]
rcustomerservice=[]
rbackroom=[]
rfloorassociate=[]

readexcelfile(path, employees, rcashier, rstocking, rcustomerservice, rbackroom,rfloorassociate)

requirement = [rcashier, rstocking, rcustomerservice, rbackroom, rfloorassociate]
categories= len(requirement)
time =  len(employees[0].availability)
training = [[employees[i].cashier, employees[i].stocking, employees[i].customerservice, employees[i].backroom, employees[i].floorassociate] for i in range(len(employees))]

print('The demands for cashiers by the company are ')
for i in range (0,len(rcashier)):
    print (rcashier[i], rstocking[i])

print('Information on the employees ')
for i in range (0,len(employees)):
    print (employees[i].eid, employees[i].pay)
    for j in range (0,len(employees[i].availability)):
        if employees[i].availability[j]==0:
            print (j, 'Not available at that time')
        if employees[i].availability[j]==1:
            print (j, 'Available at that time')


#Make your LP.
my_lp_problem = pulp.LpProblem("My LP Problem", pulp.LpMinimize)

sched = pulp.LpVariable.dicts("schedule", ((i, j, k) for i in range(len(employees)) for j in range(categories) for k in range(time)), lowBound=0, cat = 'Binary')

# obj and constr are added using += to the my_lp_problem class.
#Pulp assumes that the the objective function is alsyws given first.    
my_lp_problem += pulp.lpSum([employees[i].pay*sched[i, j, k] for i in range(len(employees)) for j in range(categories) for k in range(time)]) 

#constraints
for k in range(time):
    for j in range(categories):
        my_lp_problem += pulp.lpSum([sched[i, j, k] for i in range(len(employees))]) >= requirement[j][k]
    
for i in range(len(employees)):
    my_lp_problem += pulp.lpSum([sched[i, j, k] for j in range(categories) for k in range(time)]) <= employees[i].highhours

for i in range(len(employees)):
    my_lp_problem += pulp.lpSum([sched[i, j, k] for j in range(categories) for k in range(time)]) >= employees[i].lowhours    

for k in range(time):
    for j in range(categories):
        for i in range(len(employees)):
            my_lp_problem += sched[i, j, k] <= employees[i].availability[k]*training[i][j]
        
for k in range(time):
    for i in range(len(employees)):
        my_lp_problem += pulp.lpSum([sched[i, j, k] for j in range(categories)]) <= 1
          
# print (my_lp_problem)

status=my_lp_problem.solve()
if pulp.LpStatus[my_lp_problem.status] =='Infeasible':
    print ('INFEASIBLE ****************************************')


#To help you write out the solution.
else:
    wbr = openpyxl.Workbook()
    
    #Sheets start at 0 
    
    wbr.create_sheet(index = 0, title = "solution")
    sheetnames = wbr.sheetnames
    
    wsheet = wbr[sheetnames[0]]
    
    #Writing to individual cells
    #Cells start at 1, 1
    
    row =1 
    wsheet.cell(row,1).value ='Employee costs'
    wsheet.cell(row,2).value = pulp.value(my_lp_problem.objective)
    row+=2
    
    
    for i in range (0, len(employees)):
        wsheet.cell(row , i+2).value = employees[i].eid
        
    beginrow = row
    for k in range(time):        
        row+=1
        for i in range (0, len(employees)):
            avail = 0
            for j in range(categories):
                avail = avail + sched[i,j,k].varValue
            if avail>1:
                print("model not correct")
            wsheet.cell(row , i+2).value = avail
    
    
    days=0
    hours=10
    row=beginrow
    for i in range (0, len(rcashier)):      
        row+=1
        wsheet.cell(row , 1).value = str(days)+','+str(hours)
        hours+=1
        if hours==21:
            hours=10
            days+=1
            
    
    wbr.save('Project3Solution.xlsx')
